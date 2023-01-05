from sqlalchemy import (
    create_engine,
    MetaData
    )

from openpyxl import load_workbook

import requests
import platform
import tarfile
import pandas as pd
import datetime
import os
import sys
import yaml

def _dir_seperator_check():
	'''함수의 구분자를 os마다 판단해서 리턴하는 값'''
	if platform.system() == 'Windows':
		return '\\'
	else:
		return '/'

def _read_xlxs(base_dir, file_name):
	file_dir = base_dir + _dir_seperator_check() + file_name
	load_wb = load_workbook(file_dir, read_only=True)
	ret = load_wb['Sheet1']

	return ret

def _send_slack(url, message):
	'''slack web hook으로 text인 메세지를 보내는 함수.'''
	try:
		data = {'text' : message}
		return requests.post(url=url, json=data)
	except:
		print('webhook error!')
		print(f'detail \n{message}')

def _unzip_tar(src_dir, dst_dir, tar_name):
    with tarfile.open(src_dir + _dir_seperator_check() + tar_name, mode='r') as tar:
        try:
            tar.extractall(path=dst_dir)
        except:
            print("Error : unzip tar")

def _read_yaml(base_dir, file_name):
	file_dir = base_dir + _dir_seperator_check() + file_name
	with open(file_dir) as f:
		ret = yaml.load(f, Loader=yaml.FullLoader)
	return ret

if __name__=='__main__':
    if len(sys.argv) != 4:
        print("Error: Not Enough Argument")
        sys.exit()
    try:
        year = int(sys.argv[1])
        month = int(sys.argv[2])
        day = int(sys.argv[3])
    except:
        print("Error : Not Correct Argument")
        sys.exit()
    start_date = datetime.datetime(year=year, month=month, day=day)
    start_time = datetime.datetime.today()
    dir_seperator = _dir_seperator_check()
    dir = os.path.dirname(__file__)
    metadata_obj = MetaData()

    min_price_column = ['prdy_vrss','prdy_vrss_sign','prdy_ctrt','stck_prdy_clpr','acml_vol','acml_tr_pbmn_now','hts_kor_isnm','stck_prpr_now','stck_bsop_date','stck_cntg_hour','stck_prpr','stck_oprc','stck_hgpr','stck_lwpr','cntg_vol','acml_tr_pbmn']

    target_dir = dir + dir_seperator + 'korea_min_stock_price' + dir_seperator + start_date.strftime('%Y_%m_%d')
    kospi_dir = target_dir + dir_seperator + 'kospi'
    kosdaq_dir = target_dir + dir_seperator + 'kosdaq'

    url = _read_yaml(dir + dir_seperator + 'conf', 'url.yaml')
    try:
        _unzip_tar(target_dir, target_dir, f"{start_date.strftime('%Y_%m_%d')}.tar.gz")
    except:
        _send_slack(url['slack_webhook_url'],'Error : Not Exists Tar File')
        sys.exit()
    if not os.path.isdir(target_dir):
        _send_slack(url['slack_webhook_url'], start_date.strftime('%Y_%m_%d') + " is not exists")
        sys.exit()
    engine = create_engine("sqlite:///database.db", echo=False, future=True)

    metadata_obj.create_all(engine)
    conn = engine.connect()
    ws = _read_xlxs(kospi_dir, 'kospi_code.xlsx')
    for j in range(2, ws.max_row + 1):
        cell_num = "A" + str(j)
        cell_val = ws[cell_num].value
        try:
            df = pd.read_csv(kospi_dir + dir_seperator + cell_val + '.csv', encoding='utf-8')
            df.columns = min_price_column
            df.to_sql(name=cell_val, con=engine, if_exists='append', index=False)
        except:
            _send_slack(url['slack_webhook_url'],f"Error in {cell_val} in kospi")

    ws = _read_xlxs(kosdaq_dir, 'kosdaq_code.xlsx')
    for j in range(2, ws.max_row + 1):
        cell_num = "A" + str(j)
        cell_val = ws[cell_num].value
        try:
            df = pd.read_csv(kosdaq_dir + dir_seperator + cell_val + '.csv', encoding='utf-8')
            df.columns = min_price_column
            df.to_sql(name=cell_val, con=engine, if_exists='append', index=False)
        except:
            _send_slack(url['slack_webhook_url'],f"Error in {cell_val} in kosdaq")

    conn.close()
    total_time = datetime.datetime.today() - start_time
    print(f"total time csv to sqlite : {total_time}")
    _send_slack(url['slack_webhook_url'],f"total time csv to sqlite : {total_time}")
