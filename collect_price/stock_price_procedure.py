import yaml
import os
import requests
import json
import asyncio
import platform
import csv
import tarfile
from time import sleep

import time
import datetime
from openpyxl import load_workbook

today = datetime.datetime.today().strftime('%Y_%m_%d')

async def korea_min_stock_price(key : object, url : object, stock_num : str, csv_dir : str) -> None:
	'''
	한국투자증권 api를 이용해 한국 주식의 분봉을 수집하는 함수

	key, url 은 yaml.load를 이용해 받은 object를 권장. 이외 json도 가능

	'''
	list_stock_price = list()
	stock_price = list()
	inquire_time = datetime.datetime(year=2022, month=12, day=12, hour=9, minute=29, second=0)

	for i in range(0, 19):
		stock_price.append((await _domestic_stock_min_price(key, url, stock_num, inquire_time.strftime("%H%M%S"))).json())
		inquire_time = inquire_time + datetime.timedelta(minutes=30)
	await asyncio.sleep(0.05)
	try:
		if stock_price[0]['rt_cd'] == '0':
			await _writerow_csv(csv_dir, stock_num, list(stock_price[0]['output1'].keys()) + list((stock_price[0]['output2'])[0].keys()))
		else:
			_send_slack(url['slack_webhook_url'], "Error : Not Correct Return - " + stock_num)
	except:
		_send_slack(url['slack_webhook_url'], "Error : Not Correct Return - " + stock_num)
		return
	for i in range(0, 19):
		for output1_val in stock_price[i]['output2']:
			list_stock_price.append(list(list(stock_price[i]['output1'].values()) + list(output1_val.values())))

	list_stock_price.sort(key=lambda x:x[9])
	await _writerows_csv(csv_dir, stock_num, list_stock_price)

def kospi_stock_price_csv(base_dir, key, url, ws):
	kospi_price = "kospi"
	global today

	tmp = _create_folder(base_dir, today)
	kospi_dir = _create_folder(tmp, kospi_price)
	error_stock_num = []

	for j in range(2, ws.max_row + 1):
		cell_num = "A" + str(j)
		stock_num = ws[cell_num].value
		try:
			asyncio.run(korea_min_stock_price(key, url, stock_num, kospi_dir))
		except:
			error_stock_num.append(stock_num)
			_send_slack(url['slack_webhook_url'], "Error : " + stock_num)
			sleep(1)
			continue
	for stock_num in error_stock_num:
		try:
			asyncio.run(korea_min_stock_price(key, url, stock_num, kospi_dir))
		except:
			_send_slack(url['slack_webhook_url'], "Error again : " + stock_num)
			continue

def kosdaq_stock_price_csv(base_dir, key, url, ws):
	kosdaq_price = "kosdaq"
	global today

	tmp = _create_folder(base_dir, today)
	kosdaq_dir = _create_folder(tmp, kosdaq_price)

	for j in range(2, ws.max_row + 1):
		cell_num = "A" + str(j)
		stock_num = ws[cell_num].value
		try:
			asyncio.run(korea_min_stock_price(key, url, stock_num, kosdaq_dir))
		except:
			_send_slack(url['slack_webhook_url'], "Error : " + stock_num)
			continue

def nasdaq_stock_price(base_dir, key, url, dir_seperator):
	ws1 = _read_xlxs(base_dir + "/xlsx_file", "nas_code.xlsx")
	# ws1 = _read_xlxs(base_dir + "/xlsx_file", "nas_code.xlsx")

	# i = 1
	# start = time.time()
	# while i < 5000:
	# 	i = i + 1
	# 	cell_num = "E" + str(i)
	# 	print(cell_num)
	# 	cell_val = ws1[cell_num].value
	# 	stock_price = foreign_stock_price(key, url, "NAS", cell_val)

	# 	print((stock_price.json()["output"])["last"])
	# 	if i % 100 == 0:
	# 		end = time.time()
	# 		print(end - start)


def _read_yaml(base_dir, file_name):
	file_dir = base_dir + _dir_seperator_check() + file_name
	with open(file_dir) as f:
		ret = yaml.load(f, Loader=yaml.FullLoader)

	return ret

def _read_xlxs(base_dir, file_name):
	file_dir = base_dir + _dir_seperator_check() + file_name
	load_wb = load_workbook(file_dir, read_only=True)
	ret = load_wb['Sheet1']

	return ret

def _write_yaml(base_dir, file_name, data):
	file_dir = base_dir + _dir_seperator_check() + file_name
	with open(file_dir, 'w') as f:
		yaml.dump(data=data, stream=f)

async def _writerow_csv(base_dir, file_name, data):
	file_dir = base_dir + _dir_seperator_check() + file_name
	with open(file_dir + '.csv', 'w') as f:
		wr = csv.writer(f)
		wr.writerow(data)

async def _writerows_csv(base_dir, file_name, data):
	file_dir = base_dir + _dir_seperator_check() + file_name
	with open(file_dir + '.csv', 'a') as f:
		wr = csv.writer(f)
		wr.writerows(data)

def _new_app_token(key, url):
	app_key_url = url['real_app_domain'] + url['new_app_token']
	data = {
		"grant_type": key['grant_type'],
		"appkey": key['appkey'],
		"appsecret": key['appsecret']
	}
	ret = requests.post(url=app_key_url, data=json.dumps(data)).json()

	return ret['access_token']

def _dir_seperator_check():
	'''함수의 구분자를 os마다 판단해서 리턴하는 값'''
	if platform.system() == 'Windows':
		return '\\'
	else:
		return '/'

def _create_folder(base_dir, folder_name):
	folder = base_dir + _dir_seperator_check() + folder_name
	try:
		if not os.path.exists(folder):
			os.makedirs(folder)
	except:
		print("Error:create folder : " + folder)
	return folder

def _make_tar(folder_dir, tar_name):
	os.chdir(os.path.dirname(__file__))
	if os.path.exists(folder_dir):
		with tarfile.open(tar_name, 'w:gz') as tar:
			tar.add(folder_dir)

def _send_slack(url, message):
	'''slack web hook으로 text인 메세지를 보내는 함수.'''
	try:
		data = {'text' : message}
		return requests.post(url=url, json=data)
	except:
		print('webhook error!')
		print(f'detail \n{message}')

def _check_korea_holiday(key, url, check_day) -> requests.Response:
	header= {
		'content-type' : 'application/json; charset=utf-8',
		'authorization' : 'Bearer ' + key['apptoken'],
		'appkey' : key['appkey'],
		'appsecret': key['appsecret'],
		'tr_id' : 'CTCA0903R'
	}
	params = {
		'BASS_DT' : check_day,
		'CTX_AREA_NK' : '',
		'CTX_AREA_FK' : ''
	}
	url = url['real_app_domain'] + '/uapi/domestic-stock/v1/quotations/chk-holiday'
	try:
		ret = requests.get(url=url, headers=header, params=params)
	except:
		_send_slack(url['slack_webhook_url'], 'Error in check holiday!')
	return ret

async def _domestic_stock_price(key, url, stock_num):
	header = {
		'authorization' : "Bearer " + key['apptoken'],
		'appkey' : key['appkey'],
		'appsecret' : key['appsecret'],
		'tr_id' : 'FHKST01010100'
	}

	params = {
		'FID_COND_MRKT_DIV_CODE' : 'J',
		'FID_INPUT_ISCD' : stock_num
	}

	url = url['real_app_domain'] + '/uapi/domestic-stock/v1/quotations/inquire-price'
	try:
		ret = requests.get(url=url, headers=header, params=params)
	except:
		_send_slack(url['slack_webhook_url'], f'requests Error!\n Stock_num: {stock_num}, Time: {time}')
		pass
	return ret

async def _foreign_stock_price(key, url, excd, stock_num):
	header = {
		'authorization' : "Bearer " + key['apptoken'],
		'appkey' : key['appkey'],
		'appsecret' : key['appsecret'],
		'tr_id' : 'HHDFS00000300'
	}

	params = {
		'AUTH' : "",
		'EXCD' : excd,
		'SYMB' : stock_num
	}

	url = url['real_app_domain'] + '/uapi/overseas-price/v1/quotations/price'
	try:
		ret = requests.get(url=url, headers=header, params=params)
	except:
		_send_slack(url['slack_webhook_url'], f'requests Error!\n Stock_num: {stock_num}, Time: {time}')
		pass
	return ret

async def _domestic_stock_min_price(key, url, stock_num, time):
	header = {
		'content-type' : 'application/json; charset=utf-8',
		'authorization' : "Bearer " + key['apptoken'],
		'appkey' : key['appkey'],
		'appsecret' : key['appsecret'],
		'tr_id' : 'FHKST03010200',
		'custtype' : 'P',
	}
	params = {
		'FID_ETC_CLS_CODE' : "",
		'FID_COND_MRKT_DIV_CODE' : 'J',
		'FID_INPUT_ISCD' : stock_num,
		'FID_INPUT_HOUR_1' : time,
		'FID_PW_DATA_INCU_YN' : 'Y' #과거 데이터 포함 여부(기본은 Y, 이럴 경우 time 시간 기준 과거 시간으로 30분까지 포함하여 리턴됨.)
	}

	url = url['real_app_domain'] + '/uapi/domestic-stock/v1/quotations/inquire-time-itemchartprice'
	try:
		ret = requests.get(url=url, headers=header, params=params)
	except:
		_send_slack(url['slack_webhook_url'], f'requests Error!\n Stock_num: {stock_num}, Time: {time}')
		pass
	return ret

def main(key, url):
	base_dir = os.path.dirname(__file__)
	dir_seperator = _dir_seperator_check()
	target_dir = base_dir + dir_seperator + today

	_send_slack(url['slack_webhook_url'], 'start collect kospi min price')
	start = time.time()

	if ((_check_korea_holiday(key, url, datetime.datetime.today().strftime('%Y%m%d')).json())['output'][0]['tr_day_yn'] == 'N'):
		raise Exception

	kospi_ws = _read_xlxs(target_dir + dir_seperator + 'kospi', 'kospi_code.xlsx')
	kospi_stock_price_csv(base_dir, key, url, kospi_ws)

	end_kospi = time.time() - start
	_send_slack(url['slack_webhook_url'], f'kospi min price total time : {datetime.timedelta(seconds=end_kospi)}')

	start = time.time()
	kosdaq_ws = _read_xlxs(target_dir + dir_seperator + 'kosdaq', 'kosdaq_code.xlsx')
	kosdaq_stock_price_csv(base_dir, key, url, kosdaq_ws)

	end_kosdaq = time.time() - start
	_send_slack(url['slack_webhook_url'], f'kosdaq min price total time : {datetime.timedelta(seconds=end_kosdaq)}')

	_make_tar(today, today + '.tar.gz')
	_send_slack(url['slack_webhook_url'], 'stock min price end')
