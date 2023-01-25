from sqlalchemy import (
    create_engine,
    MetaData
    )

from openpyxl import load_workbook

import pandas as pd
import datetime
import os

import func as fun
import time


def _unzip(tar_dir, unzip_dir, tar_name):
    try:
        fun._unzip_tar(tar_dir, unzip_dir, tar_name)
    except:

        fun._send_slack(slack_url, 'Not Tar file')
        raise FileExistsError
    if not os.path.isdir(unzip_dir):
        fun._send_slack(slack_url, 'Not correct dir')
        raise FileNotFoundError

def two_table(date, db_name, slack_url):
    year, month, day = date.split(',')
    start_date = datetime.datetime(year=int(year), month=int(month), day=int(day))
    dir_sep = fun._dir_seperator_check()

    min_price_column = ['prdy_vrss','prdy_vrss_sign','prdy_ctrt','stck_prdy_clpr','acml_vol','acml_tr_pbmn_now','hts_kor_isnm','stck_prpr_now','stck_bsop_date','stck_cntg_hour','stck_prpr','stck_oprc','stck_hgpr','stck_lwpr','cntg_vol','acml_tr_pbmn']

    target_dir = os.path.dirname(__file__) + dir_sep + 'korea_min_stock_price'
    kospi_dir = target_dir + dir_sep + start_date.strftime('%Y_%m_%d') + dir_sep + 'kospi'
    kosdaq_dir = target_dir + dir_sep + start_date.strftime('%Y_%m_%d') + dir_sep + 'kosdaq'


    #_unzip(target_dir, target_dir, f"{start_date.strftime('%Y_%m_%d')}.tar.gz")

    price_table_name = 'stock_price'
    info_table_name = 'stock_info'
    engine = create_engine(f"sqlite:////{target_dir}{dir_sep}{db_name}.db", echo=False, future=True)

    metadata_obj = MetaData()

    metadata_obj.create_all(engine)
    conn = engine.connect()

    ws = fun._read_xlxs(kospi_dir, 'kospi_code.xlsx')
    day_list = []
    for j in range(2, ws.max_row + 1):
        stck_name = ws["A" + str(j)].value
        try:
            price = pd.read_csv(kospi_dir + dir_sep + stck_name + '.csv', encoding='utf-8')
            price.columns = min_price_column
            # price.rename(columns=min_price_column, inplace=True)
            day_list.append(price.iloc[0].copy(deep=True).to_list() + [stck_name])
            price['stck_name'] = stck_name
            price = price.drop(['prdy_vrss','prdy_vrss_sign','prdy_ctrt','stck_prdy_clpr','acml_vol','acml_tr_pbmn_now','hts_kor_isnm','stck_prpr_now'], axis=1)
            price.to_sql(name=price_table_name, con=engine, if_exists='append', index=True)
        except Exception as e:
            fun._send_slack(slack_url, f'Error in {stck_name} in kospi')
    day = pd.DataFrame(data=day_list)
    day.columns = min_price_column + ['stck_name']
    day = day.drop(['stck_cntg_hour','stck_prpr','stck_oprc','stck_hgpr','stck_lwpr','cntg_vol','acml_tr_pbmn'], axis=1)
    day.to_sql(name=info_table_name, con = engine, if_exists='append', index=True)

    ws = fun._read_xlxs(kosdaq_dir, 'kosdaq_code.xlsx')
    day_list = []
    for j in range(2, ws.max_row + 1):
        stck_name = ws["A" + str(j)].value
        try:
            price = pd.read_csv(kosdaq_dir + dir_sep + stck_name + '.csv', encoding='utf-8')
            price.columns = min_price_column
            # price.rename(columns=min_price_column, inplace=True)
            day_list.append(price.iloc[0].copy(deep=True).to_list() + [stck_name])
            price['stck_name'] = stck_name
            price = price.drop(['prdy_vrss','prdy_vrss_sign','prdy_ctrt','stck_prdy_clpr','acml_vol','acml_tr_pbmn_now','hts_kor_isnm','stck_prpr_now'], axis=1)
            price.to_sql(name=price_table_name, con=engine, if_exists='append', index=True)
        except Exception as e:
            fun._send_slack(slack_url, f'Error in {stck_name} in kosdaq')
    day = pd.DataFrame(data=day_list)
    day.columns = min_price_column + ['stck_name']
    day = day.drop(['stck_cntg_hour','stck_prpr','stck_oprc','stck_hgpr','stck_lwpr','cntg_vol','acml_tr_pbmn'], axis=1)
    day.to_sql(name=info_table_name, con = engine, if_exists='append', index=True)

    conn.close()


if __name__=='__main__':
    url = fun._read_yaml('conf', 'url.yaml')
    slack_url = url['slack_webhook_url']


    fun._send_slack( slack_url, f'Start stock data to database.\n {datetime.datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초 시작")}')
    for day in range(1, 25):
        year = 2023
        month = 1
        date = f"{year},{month},{day}"
        try:
            two_table(date,"two_table_2023_01", slack_url)
        except:
            fun._send_slack(slack_url, f'Error in {year},{month},{day} table.')
            continue

        fun._send_slack(slack_url, f'End {date} table \n {datetime.datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초 종료")}')
    # for day in range(1, 30):
    #     year = 2023
    #     month = 1
    #     try:
    #         diff_table(f"{year},{month},{day}","diff_table_2022_12", slack_url)
    #         #one_table(f"{year},{month},{day}","one_table_2023_01", slack_url)
    #     except:
    #         fun._send_slack(slack_url, f'Error in {year},{month},{day} table.')
    #         continue
